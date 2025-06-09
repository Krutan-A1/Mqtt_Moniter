import time
import smtplib
import csv
import json
import threading
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import paho.mqtt.client as mqtt
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# === Configuration ===
MQTT_BROKER = "192.168.2.133"
MQTT_PORT = 1883
MQTT_TOPIC = "data/sensor"
EMAIL_SENDER = "krutan.lakeshri@a1fenceproducts.com"
EMAIL_PASSWORD = "vqkfqzhbletwmqqb"  # App password
EMAIL_RECIPIENTS = ["sumit.gamre@a1fenceproducts.com"]
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
DATA_DIR = "data"
REPORT_CSV = "daily_report.csv"

REPORT_TIME_HHMM = "16:19"  # Use HH:MM format
MISSING_TIMEOUT = 30
THREAD_MESSAGE_ID = "<zaroli-monitor-thread@a1fenceproducts.com>"

# 🔧 Set to False to disable per-sensor online/offline notifications
SENSOR_LEVEL_NOTIFICATION = True
LEARNING_PERIOD = 60  # seconds to learn expected sensors
# 🔧 Set the expected number of sensors to automatically end learning phase
EXPECTED_SENSOR_COUNT = 34
# 🔧 seconds to wait before sending batched sensor status notifications
BATCH_NOTIFICATION_DELAY = 30

# === State Variables ===
last_msg_time = time.time()
data_missing_since = None
email_sent_for_missing = False
last_state_change = time.time()
broker_online_time = 0
broker_offline_time = 0
broker_last_check = time.time()
broker_was_offline = False

# Learning phase variables
start_time = time.time()
learning_phase = True
discovered_sensors = set()
EXPECTED_SENSORS = []

# mac: {last_seen, online_time, offline_time, online_status, last_notification_sent}
sensor_activity = {}

# Batched notification system
pending_notifications = []  # List of status changes to be sent in batch
last_batch_notification_time = time.time()
BATCH_NOTIFICATION_DELAY = 30  # seconds to wait before sending batch notification

# Excel workbook and worksheets for daily logging
daily_workbooks = {}  # date_str: workbook_path
sensor_data_buffer = {}  # mac: [list of data rows for current day]

# Initialize sensor tracking


def initialize_sensor(mac):
    """Initialize tracking for a new sensor"""
    current_time = time.time()
    sensor_activity[mac] = {
        "last_seen": current_time,
        "online_time": 0,
        "offline_time": 0,
        "online_status": True,
        "last_notification_sent": 0,
        "last_update": current_time
    }
    # Initialize data buffer for this sensor
    if mac not in sensor_data_buffer:
        sensor_data_buffer[mac] = []


def finalize_expected_sensors(reason="timeout"):
    """Finalize the list of expected sensors after learning period"""
    global EXPECTED_SENSORS, learning_phase

    EXPECTED_SENSORS = list(discovered_sensors)
    learning_phase = False

    print(f"🎯 Learning phase completed ({reason})!")
    print(f"📡 Discovered {len(EXPECTED_SENSORS)} sensors: {EXPECTED_SENSORS}")

    # Only send notification email if sensors were discovered or if timeout occurred
    if EXPECTED_SENSORS and reason == "timeout":
        body = (
            f"🎯 Sensor Discovery Completed!\n\n"
            f"Discovered {len(EXPECTED_SENSORS)} sensors during {LEARNING_PERIOD}-second learning period:\n\n"
        )
        for i, mac in enumerate(EXPECTED_SENSORS, 1):
            body += f"{i}. {mac}\n"

        body += f"\nMonitoring will now begin for these sensors."
        send_email(
            body, subject="Zaroli MQTT Monitor - Sensor Discovery Complete")
    elif not EXPECTED_SENSORS and reason == "timeout":
        body = (
            f"⚠️ No sensors discovered during {LEARNING_PERIOD}-second learning period.\n"
            f"Please check if sensors are transmitting data."
        )
        send_email(body, subject="Zaroli MQTT Monitor - No Sensors Discovered")
    # If reason is "count_met", don't send any email


# Ensure data directory exists
os.makedirs(DATA_DIR, exist_ok=True)

# === Functions ===


def send_email(body_text, subject="Zaroli MQTT Monitor Alert", attachment_paths=None):
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = f"Zaroli MQTT Monitor <{EMAIL_SENDER}>"
    msg["To"] = ", ".join(EMAIL_RECIPIENTS)
    msg["Message-ID"] = THREAD_MESSAGE_ID
    msg["In-Reply-To"] = THREAD_MESSAGE_ID
    msg["References"] = THREAD_MESSAGE_ID

    # Convert plain text to HTML format for better display
    html_body = f"""
    <html><body style='font-family: Arial, sans-serif;'>
    <div style='border:1px solid #ddd;padding:15px;background-color:#f9f9f9;'>
    <h3 style='color:#333;margin-top:0;'>Zaroli MQTT Monitor Notification</h3>
    <div style='white-space: pre-wrap; font-size:14px; line-height:1.4;'>{body_text}</div>
    </div>
    </body></html>
    """

    # Attach only HTML version to avoid duplication
    msg.attach(MIMEText(html_body, "html"))

    # Handle multiple attachments
    if attachment_paths:
        if isinstance(attachment_paths, str):
            attachment_paths = [attachment_paths]

        for attachment_path in attachment_paths:
            if attachment_path and os.path.exists(attachment_path):
                with open(attachment_path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header(
                        "Content-Disposition", f"attachment; filename={os.path.basename(attachment_path)}")
                    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECIPIENTS, msg.as_string())
            print("📧 Email sent")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")


def convert_to_excel_compatible(value):
    """Convert complex data types to Excel-compatible strings"""
    if isinstance(value, (list, dict, tuple)):
        return str(value)
    elif value is None:
        return ""
    elif isinstance(value, bool):
        return str(value)
    else:
        return value


def save_to_excel(data_dict):
    """Save sensor data to Excel file with separate sheets per sensor"""
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")

    mac = data_dict.get("mac")
    if not mac:
        return

    # Initialize sensor data buffer if not exists
    if mac not in sensor_data_buffer:
        sensor_data_buffer[mac] = []

    # Convert all values to Excel-compatible format
    excel_compatible_data = {}
    for key, value in data_dict.items():
        excel_compatible_data[key] = convert_to_excel_compatible(value)

    # Add timestamp and store data in buffer
    row_data = {"timestamp": timestamp, **excel_compatible_data}
    sensor_data_buffer[mac].append(row_data)

    # Write to Excel file every 10 records or every 5 minutes (whichever comes first)
    if len(sensor_data_buffer[mac]) >= 10 or should_flush_buffer(mac):
        flush_sensor_data_to_excel(mac, date_str)


def should_flush_buffer(mac):
    """Check if buffer should be flushed based on time"""
    if not sensor_data_buffer[mac]:
        return False

    # Get timestamp of oldest record in buffer
    oldest_record = sensor_data_buffer[mac][0]
    oldest_time = datetime.strptime(
        oldest_record["timestamp"], "%Y-%m-%d %H:%M:%S")

    # Flush if oldest record is more than 5 minutes old
    return (datetime.now() - oldest_time).total_seconds() > 300


def flush_sensor_data_to_excel(mac, date_str):
    """Flush sensor data buffer to Excel file"""
    if not sensor_data_buffer[mac]:
        return

    file_path = os.path.join(DATA_DIR, f"mqtt_data_{date_str}.xlsx")

    try:
        # Load existing workbook or create new one
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = Workbook()
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

        # Create or get worksheet for this sensor
        # Replace colons for valid sheet name
        sheet_name = f"Sensor_{mac.replace(':', '_')}"
        # Limit sheet name to 31 characters (Excel limit)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            start_row = worksheet.max_row + 1
        else:
            worksheet = workbook.create_sheet(sheet_name)
            # Add headers
            if sensor_data_buffer[mac]:
                headers = list(sensor_data_buffer[mac][0].keys())
                worksheet.append(headers)
            start_row = 2

        # Add buffered data to worksheet
        records_added = 0
        for row_data in sensor_data_buffer[mac]:
            try:
                # Convert all values to ensure Excel compatibility
                excel_row = []
                for value in row_data.values():
                    excel_row.append(convert_to_excel_compatible(value))
                worksheet.append(excel_row)
                records_added += 1
            except Exception as row_error:
                print(f"⚠️ Error adding row for sensor {mac}: {row_error}")
                print(f"   Problematic data: {row_data}")
                continue

        # Save workbook
        workbook.save(file_path)

        # Clear buffer for this sensor
        buffer_size = len(sensor_data_buffer[mac])
        sensor_data_buffer[mac] = []

        print(
            f"📊 Flushed {records_added}/{buffer_size} records for sensor {mac} to {file_path}")

    except Exception as e:
        print(f"❌ Error saving to Excel for sensor {mac}: {e}")
        print(f"   File: {file_path}")
        # Don't clear buffer if save failed, so we can retry later
        if "Cannot convert" in str(e):
            print(
                f"   Sample data causing error: {sensor_data_buffer[mac][:1] if sensor_data_buffer[mac] else 'No data'}")


def flush_all_sensor_buffers():
    """Flush all sensor data buffers to Excel files"""
    date_str = datetime.now().strftime("%Y-%m-%d")
    for mac in list(sensor_data_buffer.keys()):
        if sensor_data_buffer[mac]:  # Only flush if there's data
            flush_sensor_data_to_excel(mac, date_str)


def send_batch_notifications():
    """Send batched sensor status notifications"""
    global pending_notifications, last_batch_notification_time

    if not pending_notifications:
        return

    # Group notifications by type
    online_sensors = []
    offline_sensors = []

    for notification in pending_notifications:
        if notification['status'] == 'ONLINE':
            online_sensors.append(notification)
        else:
            offline_sensors.append(notification)

    # Build email body
    body_lines = [
        f"📊 Sensor Status Changes ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})",
        "",
    ]

    if online_sensors:
        body_lines.append("✅ Sensors Now ONLINE:")
        for notif in online_sensors:
            body_lines.append(
                f"  • {notif['mac']} - Back online at {notif['timestamp']}")
            if notif.get('offline_duration', 0) > 0:
                body_lines.append(
                    f"    (Was offline for {notif['offline_duration']} seconds)")
        body_lines.append("")

    if offline_sensors:
        body_lines.append("🚨 Sensors Now OFFLINE:")
        for notif in offline_sensors:
            body_lines.append(
                f"  • {notif['mac']} - Went offline at {notif['timestamp']}")
            body_lines.append(
                f"    (No data for {notif['silent_duration']} seconds)")
        body_lines.append("")

    # Add summary
    total_online = len(online_sensors)
    total_offline = len(offline_sensors)
    body_lines.append(
        f"📈 Summary: {total_online} sensors came online, {total_offline} sensors went offline")

    # Send the batched notification
    subject = "Sensor Status Changes" + (f" - {total_online} Online" if total_online > 0 else "") + (
        f", {total_offline} Offline" if total_offline > 0 else "")
    send_email("\n".join(body_lines), subject=subject)

    # Clear pending notifications
    pending_notifications = []
    last_batch_notification_time = time.time()


def check_sensor_status():
    """Check all expected sensors and queue notifications for status changes"""
    global pending_notifications, last_batch_notification_time

    # Skip sensor monitoring during learning phase
    if learning_phase:
        return

    current_time = time.time()

    for mac in EXPECTED_SENSORS:
        if mac not in sensor_activity:
            continue

        activity = sensor_activity[mac]
        time_since_last_seen = current_time - \
            activity["last_seen"] if activity["last_seen"] > 0 else float(
                'inf')
        currently_online = time_since_last_seen <= MISSING_TIMEOUT and activity[
            "last_seen"] > 0
        previously_online = activity["online_status"]

        # Update uptime/downtime
        if activity["last_seen"] > 0:
            duration = current_time - \
                max(activity["last_seen"], activity.get(
                    "last_update", activity["last_seen"]))
            if previously_online:
                activity["online_time"] += duration
            else:
                activity["offline_time"] += duration

        activity["last_update"] = current_time

        # Check for status change and queue notification if enabled
        if SENSOR_LEVEL_NOTIFICATION and currently_online != previously_online:
            # Avoid duplicate notifications within 60 seconds
            if current_time - activity.get("last_notification_sent", 0) > 60:
                activity["online_status"] = currently_online
                activity["last_notification_sent"] = current_time

                # Queue notification instead of sending immediately
                notification = {
                    'mac': mac,
                    'status': 'ONLINE' if currently_online else 'OFFLINE',
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                }

                if currently_online:
                    offline_duration = int(
                        time_since_last_seen) if time_since_last_seen < float('inf') else 0
                    notification['offline_duration'] = offline_duration
                else:
                    notification['silent_duration'] = int(time_since_last_seen)

                pending_notifications.append(notification)

        activity["online_status"] = currently_online

    # Send batch notification if enough time has passed and there are pending notifications
    if (pending_notifications and
            current_time - last_batch_notification_time >= BATCH_NOTIFICATION_DELAY):
        send_batch_notifications()


def on_connect(client, userdata, flags, rc):
    if rc == 0:
        print("✅ Connected to MQTT broker")
        client.subscribe(MQTT_TOPIC)
    else:
        print(f"❌ Connection failed with code {rc}")


def on_message(client, userdata, msg):
    global last_msg_time, data_missing_since, email_sent_for_missing, last_state_change, broker_was_offline
    global learning_phase, discovered_sensors

    now = time.time()
    last_msg_time = now

    try:
        payload = json.loads(msg.payload.decode())
        mac = payload.get("mac")

        # Save to Excel with separate sheets per sensor
        save_to_excel(payload)

        if mac:
            # During learning phase, collect discovered sensors
            if learning_phase:
                if mac not in discovered_sensors:
                    discovered_sensors.add(mac)
                    print(
                        f"📡 Discovered sensor: {mac} ({len(discovered_sensors)} total)")

                    # Check if we've reached the expected sensor count
                    if len(discovered_sensors) >= EXPECTED_SENSOR_COUNT:
                        print(
                            f"🎯 Expected sensor count ({EXPECTED_SENSOR_COUNT}) reached!")
                        finalize_expected_sensors(reason="count_met")

                # Initialize sensor tracking even during learning
                if mac not in sensor_activity:
                    initialize_sensor(mac)
                else:
                    sensor_activity[mac]["last_seen"] = now
                    sensor_activity[mac]["last_update"] = now
            else:
                # Normal operation after learning phase
                if mac not in sensor_activity:
                    # Unknown sensor (not discovered during learning phase)
                    initialize_sensor(mac)
                    print(
                        f"⚠️ Unknown sensor detected: {mac} (not in expected list)")
                else:
                    activity = sensor_activity[mac]

                    # Update uptime for the gap since last seen
                    if activity["last_seen"] > 0:
                        duration = now - \
                            activity.get("last_update", activity["last_seen"])
                        if activity["online_status"]:
                            activity["online_time"] += duration
                        else:
                            activity["offline_time"] += duration

                    activity["last_seen"] = now
                    activity["last_update"] = now

    except Exception as e:
        print(f"⚠️ Error parsing message: {e}")

    if email_sent_for_missing:
        down_duration = int(now - data_missing_since)
        last_state_change = now
        broker_was_offline = False
        body = (
            f"✅ Data resumed on topic '{MQTT_TOPIC}' at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"Data was missing for {down_duration} seconds."
        )
        send_email(body)
        email_sent_for_missing = False
        data_missing_since = None


def create_daily_report():
    """Create daily report CSV file"""
    current_timestamp = time.time()
    online_sensors = 0
    offline_sensors = []
    unknown_sensors = []

    report_file_path = os.path.join(DATA_DIR, REPORT_CSV)

    with open(report_file_path, mode="w", newline="") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["MAC Address", "Status", "Online Time (s)",
                        "Offline Time (s)", "Last Seen", "Expected"])

        # Process expected sensors (only if learning phase is complete)
        if not learning_phase and EXPECTED_SENSORS:
            for mac in EXPECTED_SENSORS:
                if mac in sensor_activity:
                    data = sensor_activity[mac]

                    # Update final duration before report
                    if data["last_seen"] > 0:
                        duration = current_timestamp - \
                            data.get("last_update", data["last_seen"])
                        time_since_last_seen = current_timestamp - \
                            data["last_seen"]

                        if time_since_last_seen <= MISSING_TIMEOUT:
                            data["online_time"] += duration
                            data["online_status"] = True
                            online_sensors += 1
                            status = "ONLINE"
                        else:
                            data["offline_time"] += duration
                            data["online_status"] = False
                            offline_sensors.append(mac)
                            status = "OFFLINE"

                        last_seen_str = datetime.fromtimestamp(
                            data["last_seen"]).strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        offline_sensors.append(mac)
                        status = "NEVER SEEN"
                        last_seen_str = "Never"

                    writer.writerow([mac, status, int(data['online_time']), int(
                        data['offline_time']), last_seen_str, "Yes"])
                else:
                    # Sensor never seen
                    offline_sensors.append(mac)
                    # 24 hours offline if never seen
                    writer.writerow(
                        [mac, "NEVER SEEN", 0, 86400, "Never", "Yes"])

        # Process all sensors (including unexpected ones)
        for mac, data in sensor_activity.items():
            if not learning_phase and mac not in EXPECTED_SENSORS:
                unknown_sensors.append(mac)
                expected_status = "No"
            elif learning_phase:
                expected_status = "Learning"
            else:
                continue  # Skip expected sensors as they're already processed above

            duration = current_timestamp - \
                data.get("last_update", data["last_seen"])
            time_since_last_seen = current_timestamp - data["last_seen"]

            if time_since_last_seen <= MISSING_TIMEOUT:
                data["online_time"] += duration
                status = "ONLINE"
                if learning_phase:
                    online_sensors += 1
            else:
                data["offline_time"] += duration
                status = "OFFLINE"
                if learning_phase:
                    offline_sensors.append(mac)

            last_seen_str = datetime.fromtimestamp(
                data["last_seen"]).strftime('%Y-%m-%d %H:%M:%S')
            writer.writerow([mac, status, int(data['online_time']), int(
                data['offline_time']), last_seen_str, expected_status])

    return report_file_path, online_sensors, offline_sensors, unknown_sensors


def report_daily():
    global sensor_activity, broker_online_time, broker_offline_time, broker_last_check

    while True:
        now = datetime.now()
        report_time = datetime.strptime(REPORT_TIME_HHMM, "%H:%M").time()
        next_report = datetime.combine(now.date(), report_time)
        if now.time() > report_time:
            next_report += timedelta(days=1)

        sleep_duration = (next_report - now).total_seconds()
        time.sleep(sleep_duration)

        # Flush all remaining sensor data before creating report
        flush_all_sensor_buffers()

        # Create daily report
        report_file_path, online_sensors, offline_sensors, unknown_sensors = create_daily_report()

        report_lines = [
            f"📊 Daily Report ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}):",
            "",
            f"🔧 Broker Statistics:",
            f"  ✅ Online Time: {int(broker_online_time)} seconds ({broker_online_time/3600:.1f} hours)",
            f"  ❌ Offline Time: {int(broker_offline_time)} seconds ({broker_offline_time/3600:.1f} hours)",
            "",
        ]

        if learning_phase:
            report_lines.extend([
                f"🎯 System Status: Still in learning phase",
                f"📡 Sensors discovered so far: {len(discovered_sensors)}",
                f"⏱️ Learning period: {LEARNING_PERIOD} seconds",
                ""
            ])
        else:
            report_lines.extend([
                f"📡 Sensor Summary:",
                f"  ✅ Online sensors: {online_sensors}/{len(EXPECTED_SENSORS)}",
                f"  ❌ Offline/Missing sensors: {len(offline_sensors)}/{len(EXPECTED_SENSORS)}",
                ""
            ])

        if offline_sensors:
            if learning_phase:
                report_lines.append("⚠️ Currently Offline Sensors:")
            else:
                report_lines.append("🚨 Offline/Missing Expected Sensors:")

            for mac in offline_sensors:
                if mac in sensor_activity and sensor_activity[mac]["last_seen"] > 0:
                    last_seen = datetime.fromtimestamp(
                        sensor_activity[mac]["last_seen"]).strftime('%Y-%m-%d %H:%M:%S')
                    report_lines.append(f"  - {mac} (Last seen: {last_seen})")
                else:
                    report_lines.append(f"  - {mac} (Never seen)")
            report_lines.append("")

        if unknown_sensors:
            report_lines.append(
                "ℹ️ Unexpected Sensors (not in expected list):")
            for mac in unknown_sensors:
                last_seen = datetime.fromtimestamp(
                    sensor_activity[mac]["last_seen"]).strftime('%Y-%m-%d %H:%M:%S')
                status = "ONLINE" if time.time(
                ) - sensor_activity[mac]["last_seen"] <= MISSING_TIMEOUT else "OFFLINE"
                report_lines.append(
                    f"  - {mac} ({status}, Last seen: {last_seen})")
            report_lines.append("")

        # Get today's Excel data file
        today_file = os.path.join(
            DATA_DIR, f"mqtt_data_{datetime.now().strftime('%Y-%m-%d')}.xlsx")

        # Prepare attachments
        attachments = [report_file_path]
        if os.path.exists(today_file):
            attachments.append(today_file)
            report_lines.append(
                "📎 Attachments: Daily sensor report CSV + MQTT data Excel file")
            report_lines.append(
                "📊 Excel file contains separate sheets for each sensor")
        else:
            report_lines.append("📎 Attachments: Daily sensor report CSV")
            report_lines.append("⚠️ No MQTT data file found for today")

        # Send combined email with both attachments and report in body
        send_email(
            "\n".join(report_lines),
            subject="Zaroli MQTT Monitor - Daily Summary & Data",
            attachment_paths=attachments
        )

        # Reset counters for next day
        for data in sensor_activity.values():
            data["online_time"] = 0
            data["offline_time"] = 0
            data["last_update"] = time.time()
        broker_online_time = 0
        broker_offline_time = 0
        broker_last_check = time.time()


def periodic_flush():
    """Periodically flush sensor data buffers"""
    while True:
        time.sleep(60)  # Flush every minute
        flush_all_sensor_buffers()


# MQTT Setup
client = mqtt.Client()
client.on_connect = on_connect
client.on_message = on_message
client.connect(MQTT_BROKER, MQTT_PORT, 60)
client.loop_start()

# Start report thread
threading.Thread(target=report_daily, daemon=True).start()

# Start periodic flush thread
threading.Thread(target=periodic_flush, daemon=True).start()

print(f"🎯 Starting sensor discovery phase...")
print(
    f"📡 Looking for {EXPECTED_SENSOR_COUNT} sensors (max {LEARNING_PERIOD} seconds)...")

# Monitor loop
try:
    while True:
        current_time = time.time()

        # Check if learning phase should end due to timeout
        if learning_phase and (current_time - start_time) >= LEARNING_PERIOD:
            finalize_expected_sensors(reason="timeout")

        time_since_last = current_time - last_msg_time

        # Update broker uptime/downtime
        time_diff = current_time - broker_last_check
        if time_since_last <= MISSING_TIMEOUT:
            broker_online_time += time_diff
            broker_was_offline = False
        else:
            broker_offline_time += time_diff
            broker_was_offline = True

        broker_last_check = current_time

        # Check individual sensor status (only after learning phase)
        if not learning_phase:
            check_sensor_status()

            # Send any remaining pending notifications at the end of the day
            # 5 minutes
            if pending_notifications and (current_time - last_batch_notification_time) >= 300:
                send_batch_notifications()

        # Check for broker-level data missing (only after learning phase)
        if not learning_phase and time_since_last > MISSING_TIMEOUT and not email_sent_for_missing:
            data_missing_since = last_msg_time
            last_state_change = current_time
            broker_was_offline = True
            body = (
                f"🚨 No data received on topic '{MQTT_TOPIC}' since {datetime.fromtimestamp(last_msg_time).strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"This alert is sent after {MISSING_TIMEOUT} seconds of silence."
            )
            send_email(body, subject="MQTT Broker - No Data Alert")
            email_sent_for_missing = True

        time.sleep(5)

except KeyboardInterrupt:
    print("🛑 Monitoring stopped by user.")
    # Flush all remaining data before shutdown
    flush_all_sensor_buffers()
    client.loop_stop()
    client.disconnect()
